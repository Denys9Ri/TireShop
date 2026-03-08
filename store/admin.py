from django.contrib import admin
from django.urls import path
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django import forms
from django.http import HttpResponse, FileResponse
from django.utils.html import format_html
from django.template.loader import render_to_string
from django.db.models import Q 
from django.conf import settings
import openpyxl
import re
import gc
import io
import os
import tempfile
import urllib.request
import traceback
from xhtml2pdf import pisa
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from .models import Product, Brand, Order, OrderItem, ProductImage, SiteSettings, AboutImage

# ==========================================
# 🔥 ЛОГІКА ГЕНЕРАЦІЇ PDF-ЧЕКА (КАТАЛОГ + ШРИФТИ) 🔥
# ==========================================

def get_cyrillic_font_path():
    sys_fonts = [
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
        '/usr/share/fonts/truetype/freefont/FreeSans.ttf'
    ]
    for f in sys_fonts:
        if os.path.exists(f):
            return f
            
    font_path = os.path.join(tempfile.gettempdir(), 'Roboto-Regular.ttf')
    if not os.path.exists(font_path):
        try:
            url = "https://github.com/googlefonts/roboto/raw/main/src/hinted/Roboto-Regular.ttf"
            urllib.request.urlretrieve(url, font_path)
        except Exception as e:
            print("Помилка завантаження шрифту:", e)
            
    return font_path if os.path.exists(font_path) else None

def link_callback(uri, rel):
    if uri.startswith('file://'):
        return uri.replace('file://', '')
    return uri

def generate_order_pdf(order):
    font_path = get_cyrillic_font_path()
    font_uri = ""
    
    if font_path:
        try:
            pdfmetrics.registerFont(TTFont('CyrillicFont', font_path))
        except:
            pass
        clean_path = font_path.replace('\\', '/')
        if not clean_path.startswith('/'):
            clean_path = '/' + clean_path
        font_uri = 'file://' + clean_path

    total_cost = sum(item.get_cost() for item in order.items.all())
    
    context = {
        'order': order, 
        'total_cost': total_cost,
        'font_uri': font_uri 
    }
    
    html_string = render_to_string('store/pdf/invoice.html', context)
    result = io.BytesIO()
    
    pisa_status = pisa.CreatePDF(
        io.BytesIO(html_string.encode("UTF-8")), 
        dest=result, 
        encoding='UTF-8',
        link_callback=link_callback 
    )
    
    if pisa_status.err:
        return None
    result.seek(0)
    return result

# --- ЗАМОВЛЕННЯ ---
class OrderItemInline(admin.TabularInline):
    model = OrderItem
    raw_id_fields = ['product']
    extra = 0
    fields = ['product', 'quantity', 'price_at_purchase', 'get_cost_display']
    readonly_fields = ['get_cost_display']
    
    def get_cost_display(self, obj):
        return f"{obj.get_cost():.2f} грн"
    get_cost_display.short_description = "Сума"

@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display = ['id', 'status', 'created_at', 'full_name', 'phone', 'shipping_type', 'order_items_summary', 'total_cost', 'print_invoice_button']
    list_filter = ['status', 'created_at', 'shipping_type']
    search_fields = ['id', 'full_name', 'phone', 'email']
    inlines = [OrderItemInline]
    list_editable = ['status']
    readonly_fields = ['created_at', 'total_cost_detailed']

    def get_changelist_form(self, request, **kwargs):
        form = super().get_changelist_form(request, **kwargs)
        if 'status' in form.base_fields:
            form.base_fields['status'].widget.attrs['style'] = 'min-width: 170px;'
        return form

    fieldsets = (
        ('Статус та Доставка', {'fields': ('status', 'shipping_type', 'created_at')}),
        ('Дані Клієнта', {'fields': ('full_name', 'phone', 'email')}),
        ('Адреса (для НП)', {'fields': ('city', 'nova_poshta_branch')}),
        ('Фінанси', {'fields': ('total_cost_detailed',)}),
    )

    def total_cost(self, obj):
        sum_val = sum(item.get_cost() for item in obj.items.all())
        return format_html("<b>{} грн</b>", f"{sum_val:.2f}")
    total_cost.short_description = 'Сума'

    def total_cost_detailed(self, obj):
        sum_val = sum(item.get_cost() for item in obj.items.all())
        return f"{sum_val:.2f} грн"
    total_cost_detailed.short_description = 'Разом до сплати'

    def order_items_summary(self, obj):
        items = obj.items.all()
        result = []
        for item in items:
            if item.product:
                brand = item.product.brand.name if item.product.brand else "Без бренду"
                name = item.product.display_name 
                row = f"<span style='color:#0d6efd;'><b>{brand}</b></span> &nbsp;|&nbsp; <b>{item.quantity} шт.</b><br><span style='font-size:0.9em; color:#555;'>{name}</span>"
                result.append(row)
            else:
                result.append(f"Видалений товар ({item.quantity} шт.)")
        return format_html("<hr style='margin:5px 0;'>".join(result))
    order_items_summary.short_description = 'Що замовили'

    def print_invoice_button(self, obj):
        return format_html(
            '<a class="button" href="{}/print/" target="_blank" style="background-color: #0d6efd; padding: 5px 10px; border-radius: 4px; color: white; white-space: nowrap;">📄 Чек</a>',
            obj.id
        )
    print_invoice_button.short_description = 'Друк'

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [path('<int:order_id>/print/', self.admin_print_invoice, name="order_print_invoice")]
        return my_urls + urls

    def admin_print_invoice(self, request, order_id):
        try:
            order = get_object_or_404(Order, id=order_id)
            pdf_file = generate_order_pdf(order)
            
            if pdf_file:
                filename = f"check_R16_{order.id}.pdf"
                return FileResponse(pdf_file, as_attachment=False, content_type='application/pdf', filename=filename)
            else:
                return HttpResponse("Помилка: Бібліотека PDF повернула пустий файл.")
        except Exception as e:
            error_trace = traceback.format_exc()
            return HttpResponse(f"<h2 style='color:red;'>Помилка PDF</h2><pre>{error_trace}</pre>")

# --- ІНШІ НАЛАШТУВАННЯ АДМІНКИ ---
class ProductImageInline(admin.TabularInline):
    model = ProductImage
    extra = 1
    fields = ('image_url', 'image', 'preview')
    readonly_fields = ('preview',)
    def preview(self, obj):
        if obj.image_url: return format_html('<img src="{}" style="height: 50px; border-radius: 4px;"/>', obj.image_url)
        if obj.image: return format_html('<img src="{}" style="height: 50px; border-radius: 4px;"/>', obj.image.url)
        return "-"

@admin.register(SiteSettings)
class SiteSettingsAdmin(admin.ModelAdmin):
    list_display = ['global_markup']
    def has_add_permission(self, request): return not SiteSettings.objects.exists()

# --- ФОРМИ ІМПОРТУ ---
class ExcelImportForm(forms.Form):
    excel_file = forms.FileField(label="Прайс-лист (Товари)")
    start_row = forms.IntegerField(initial=2, min_value=2, label="Почати з рядка")
    end_row = forms.IntegerField(initial=2000, min_value=2, label="Закінчити рядком")

class PhotoImportForm(forms.Form):
    excel_file = forms.FileField(label="Файл з ФОТО (Brand, Model, URL)")

class SeoImportForm(forms.Form):
    excel_file = forms.FileField(label="SEO Файл (.xlsx)")
    start_row = forms.IntegerField(initial=2, min_value=2, label="Почати з рядка")
    end_row = forms.IntegerField(initial=500, min_value=2, label="Закінчити рядком")


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ['name', 'brand', 'width', 'profile', 'diameter', 'price_display', 'discount_percent', 'stock_quantity', 'photo_preview']
    list_filter = ['brand', 'seasonality', 'diameter', 'stud_type']
    search_fields = ['name', 'width', 'brand__name', 'slug']
    change_list_template = "store/admin_changelist.html"
    readonly_fields = ["photo_preview", "final_price_preview"]
    inlines = [ProductImageInline]
    prepopulated_fields = {'slug': ('name',)}
    fieldsets = (
        (None, {'fields': ('name', 'slug', 'brand', 'width', 'profile', 'diameter', 'seasonality', 'description')}),
        ('SEO', {'fields': ('seo_title', 'seo_h1', 'seo_text')}), 
        ('Фінанси', {'fields': ('cost_price', 'discount_percent', 'final_price_preview', 'stock_quantity')}),
        ('Фото', {'fields': ('photo', 'photo_url', 'photo_preview')}),
        ('Хар-ки', {'fields': ('country', 'year', 'load_index', 'speed_index', 'stud_type', 'vehicle_type')}),
    )
    def price_display(self, obj): return f"{obj.price} грн"
    price_display.short_description = "Ціна"
    def final_price_preview(self, obj): return f"{obj.price} грн"
    def photo_preview(self, obj):
        if obj.photo_url: return format_html('<img src="{}" style="max-height: 50px;"/>', obj.photo_url)
        return "—"
    photo_preview.short_description = "Фото"
    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-excel/', self.import_excel, name="import_excel"),
            path('import-photos/', self.import_photos, name="import_photos"),
            path('import-seo/', self.import_seo, name="import_seo"),
            path('export-models/', self.export_unique_models, name="export_unique_models"),
        ]
        return my_urls + urls

    # 🔥 ПОВЕРНУВ ТВОЇ ФУНКЦІЇ ЕКСПОРТУ/ІМПОРТУ 🔥
    def export_unique_models(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Models"
        ws.append(['Brand', 'Clean Model Name', 'Photo URL'])
        
        products = Product.objects.all().select_related('brand')
        seen = set()
        
        for p in products:
            brand_name = p.brand.name if p.brand else "Unknown"
            raw_name = p.name
            
            clean = re.sub(r'шина', '', raw_name, flags=re.IGNORECASE)
            clean = re.sub(r'\b\d{3}/\d{2}R?\d{0,2}\b', '', clean) 
            clean = re.sub(r'\bR\d{2}C?\b', '', clean)
            clean = re.sub(r'\b\d{2,3}[A-Z]\b', '', clean)
            if p.brand:
                clean = re.sub(rf'\({re.escape(p.brand.name)}\)', '', clean, flags=re.IGNORECASE)
                clean = re.sub(rf'\b{re.escape(p.brand.name)}\b', '', clean, flags=re.IGNORECASE)
            
            clean = clean.replace('()', '').strip()
            clean = re.sub(r'\s+', ' ', clean)
            
            if len(clean) < 2: clean = p.name

            key = (brand_name.upper(), clean.upper())
            if key not in seen:
                ws.append([brand_name, clean, ''])
                seen.add(key)
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=clean_models_for_photos.xlsx'
        wb.save(response)
        return response

    def import_photos(self, request):
        if request.method == "POST":
            excel_file = request.FILES["excel_file"]
            try:
                wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
                sheet = wb.active
                updated_products = 0
                
                IGNORE_WORDS = [
                    'serbia', 'china', 'korea', 'thailand', 'japan', 'turkey', 'germany', 'poland', 
                    'dot', 'xl', 'new', 'demo', 'usa', 'hungary', 'romania', 'france', 'spain'
                ]

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not row or len(row) < 3: continue
                    if not row[0] or not row[1] or not row[2]: continue
                    
                    brand_txt = str(row[0]).strip()
                    model_txt = str(row[1]).strip()
                    url_txt = str(row[2]).strip()
                    
                    if not url_txt.startswith('http'): continue

                    query = Q(brand__name__icontains=brand_txt)
                    
                    clean_model_txt = re.sub(r'[(),]', ' ', model_txt)
                    model_tokens = clean_model_txt.split()
                    
                    valid_tokens = []
                    for token in model_tokens:
                        t_lower = token.lower()
                        if len(token) > 1 and t_lower not in IGNORE_WORDS:
                            valid_tokens.append(token)
                    
                    if not valid_tokens: continue

                    for token in valid_tokens:
                        query &= Q(name__icontains=token)

                    count = Product.objects.filter(query).update(photo_url=url_txt)
                    updated_products += count
                
                messages.success(request, f"✅ Успіх! Фото оновлено для {updated_products} товарів!")
            except Exception as e:
                messages.error(request, f"Помилка: {e}")
            return redirect("..")
            
        form = PhotoImportForm()
        return render(request, "store/admin_import_photos.html", {"form": form})

    def import_excel(self, request):
        if request.method == "POST":
            form = ExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = form.cleaned_data["excel_file"]
                start_row_limit = form.cleaned_data["start_row"]
                end_row_limit = form.cleaned_data["end_row"]
                
                try:
                    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
                    sheet = wb.active
                    created_count = 0; updated_count = 0
                    existing_brands = {b.name.upper(): b for b in Brand.objects.all()}
                    rows_iter = sheet.iter_rows(values_only=True)
                    try: header_row = next(rows_iter)
                    except: return redirect("..")

                    def find_col(aliases):
                        for idx, cell in enumerate(header_row):
                            val = str(cell or "").strip().lower()
                            for alias in aliases:
                                if val.startswith(alias): return idx
                        return None

                    c_brand = find_col(["бренд", "brand"]) or 0
                    c_model = find_col(["модель", "model"]) or 1
                    c_size = find_col(["типоразмер", "size"]) or 2
                    c_season = find_col(["сезон", "season"]) or 3
                    c_price = find_col(["цена", "price"]) or 4
                    c_qty = find_col(["кол", "qty"]) or 5
                    c_country = find_col(["країна", "country"]) or 6
                    c_year = find_col(["рік", "year"]) or 7
                    c_photo = find_col(["фото", "photo"])
                    
                    for i, row in enumerate(rows_iter):
                        current_excel_row = i + 2
                        if current_excel_row < start_row_limit: continue
                        if current_excel_row > end_row_limit: break
                        if i % 100 == 0: gc.collect()
                        
                        if not row or len(row) < 2: continue
                        if not row[c_brand] and not row[c_model]: continue

                        brand_name = str(row[c_brand]).strip()
                        if not brand_name or brand_name == "None": brand_name = "Unknown"
                        brand_key = brand_name.upper()
                        if brand_key in existing_brands: brand_obj = existing_brands[brand_key]
                        else:
                            brand_obj = Brand.objects.create(name=brand_name)
                            existing_brands[brand_key] = brand_obj

                        model_name = str(row[c_model]).strip()
                        size_raw = str(row[c_size]).strip()
                        match = re.search(r'(\d+)/(\d+)\s*[a-zA-Z]*\s*(\d+)', size_raw)
                        if match: w, p, d = int(match.group(1)), int(match.group(2)), int(match.group(3))
                        else: w, p, d = 0, 0, 0

                        unique_name = model_name
                        if (w == 0 or p == 0 or d == 0) and size_raw: unique_name = f"{model_name} [{size_raw}]"

                        season_raw = str(row[c_season]).lower() if row[c_season] else ""
                        season_key = 'all-season'
                        if 'зим' in season_raw or 'winter' in season_raw: season_key = 'winter'
                        elif 'літ' in season_raw or 'summer' in season_raw: season_key = 'summer'

                        try:
                            raw_price = row[c_price]
                            if isinstance(raw_price, (int, float)): cost = float(raw_price)
                            else:
                                clean_price = re.sub(r'[^\d,.]', '', str(raw_price)).replace(',', '.')
                                if clean_price.count('.') > 1:
                                    parts = clean_price.split('.')
                                    clean_price = "".join(parts[:-1]) + "." + parts[-1]
                                cost = float(clean_price)
                        except: cost = 0.0

                        try:
                            qty_val = str(row[c_qty]).strip()
                            if '>' in qty_val: qty = 20
                            else: qty = int(re.sub(r'[^0-9]', '', qty_val) or 0)
                        except: qty = 0

                        country = str(row[c_country]).strip() if c_country and row[c_country] else "-"
                        try: year = int(row[c_year]) if c_year and row[c_year] else 2024
                        except: year = 2024
                        
                        photo_link = str(row[c_photo]).strip() if c_photo and row[c_photo] else None

                        obj, created = Product.objects.update_or_create(
                            name=unique_name, brand=brand_obj, width=w, profile=p, diameter=d,
                            defaults={
                                'seasonality': season_key, 'cost_price': cost, 'stock_quantity': qty,
                                'country': country, 'year': year,
                                'description': f"Шини {brand_name} {model_name}. {size_raw}. {season_raw}."
                            }
                        )
                        if photo_link and not obj.photo_url:
                            obj.photo_url = photo_link
                            obj.save(update_fields=['photo_url'])

                        if created: created_count += 1
                        else: updated_count += 1

                    messages.success(request, f"Частина оброблена! ({start_row_limit}-{end_row_limit}). ✅: {created_count}, 🔄: {updated_count}")
                except Exception as e: messages.error(request, f"Помилка: {e}")
                return redirect("..")
        else: form = ExcelImportForm()
        return render(request, "store/admin_import.html", {"form": form})

    def import_seo(self, request):
        if request.method == "POST":
            form = SeoImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = form.cleaned_data["excel_file"]
                start_row_limit = form.cleaned_data["start_row"]
                end_row_limit = form.cleaned_data["end_row"]

                try:
                    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
                    sheet = wb.active
                    updated_count = 0
                    not_found_count = 0
                    
                    rows_iter = sheet.iter_rows(values_only=True)
                    
                    try:
                        header_row = next(rows_iter)
                        header = [str(h).lower().strip() for h in header_row]
                    except:
                        messages.error(request, "Файл пустий")
                        return redirect("..")

                    try:
                        idx_brand = header.index('brand')
                        idx_model = header.index('model')
                        idx_title = header.index('title')
                        idx_h1 = header.index('h1')
                        idx_text = header.index('seo text')
                    except ValueError as e:
                        messages.error(request, f"Не знайдено колонку: {e}")
                        return redirect("..")

                    for i, row in enumerate(rows_iter):
                        current_excel_row = i + 2
                        if current_excel_row < start_row_limit: continue
                        if current_excel_row > end_row_limit: break
                        if i % 100 == 0: gc.collect()

                        if not row or len(row) < 2: continue
                        
                        brand_val = str(row[idx_brand]).strip()
                        model_val = str(row[idx_model]).strip()
                        
                        if not brand_val or not model_val: continue

                        seo_title = str(row[idx_title]).strip() if row[idx_title] else ""
                        seo_h1 = str(row[idx_h1]).strip() if row[idx_h1] else ""
                        seo_text = str(row[idx_text]).strip() if row[idx_text] else ""

                        query = Q(brand__name__icontains=brand_val)
                        for token in model_val.split():
                             if len(token) > 1: query &= Q(name__icontains=token)

                        product = Product.objects.filter(query).first()
                        
                        if product:
                            product.seo_title = seo_title
                            product.seo_h1 = seo_h1
                            product.seo_text = seo_text
                            product.save()
                            updated_count += 1
                        else:
                            not_found_count += 1

                    messages.success(request, f"✅ Оновлено SEO: {updated_count}. ❌ Не знайдено: {not_found_count}.")

                except Exception as e:
                    messages.error(request, f"Помилка: {e}")
                return redirect("..")
        else:
            form = SeoImportForm()
        
        return render(request, "store/admin_import.html", {"form": form, "title": "Імпорт SEO"})

@admin.register(Brand)
class BrandAdmin(admin.ModelAdmin):
    list_display = ['name', 'category', 'country'] 
    list_editable = ['category'] 
    list_filter = ['category']
    search_fields = ['name']

@admin.register(AboutImage)
class AboutImageAdmin(admin.ModelAdmin):
    list_display = ['id', 'created_at', 'image_preview']
    ordering = ['-created_at']
    def image_preview(self, obj):
        if obj.image: return format_html('<img src="{}" style="height: 100px; border-radius: 4px;"/>', obj.image.url)
        return "-"
    image_preview.short_description = "Попередній перегляд"
